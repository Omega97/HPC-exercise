from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "HPC_numeric_results.xlsx"

DATASETS = [
    (
        "Ex1_Bcast_EPYC",
        ROOT / "exercise1/results/bcastEPYC.csv",
        "Exercise 1 — MPI_Bcast latency on EPYC (Orfeo). Multiple measurement runs are stacked.",
    ),
    (
        "Ex1_Bcast_THIN",
        ROOT / "exercise1/results/bcastTHIN.csv",
        "Exercise 1 — MPI_Bcast latency on THIN nodes (message size sweep).",
    ),
    (
        "Ex1_Barrier_EPYC",
        ROOT / "exercise1/results/barrierEPYC.csv",
        "Exercise 1 — MPI_Barrier latency on EPYC.",
    ),
    (
        "Ex1_Barrier_THIN",
        ROOT / "exercise1/results/barrierTHIN.csv",
        "Exercise 1 — MPI_Barrier latency on THIN nodes.",
    ),
    (
        "Ex2_Strong_OMP_EPYC",
        ROOT / "exercise2/results/strong_scaling_OMP_EPYC.csv",
        "Exercise 2 — Mandelbrot strong scaling (OpenMP, EPYC).",
    ),
    (
        "Ex2_Weak_OMP_EPYC",
        ROOT / "exercise2/results/weak_scaling_OMP_EPYC.csv",
        "Exercise 2 — Mandelbrot weak scaling (OpenMP, EPYC).",
    ),
]

index_rows = []
with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    for sheet, path, desc in DATASETS:
        df = pd.read_csv(path)
        df.to_excel(writer, sheet_name=sheet, index=False)
        index_rows.append(
            {
                "Sheet": sheet,
                "Source file": path.relative_to(ROOT).as_posix(),
                "Rows": len(df),
                "Columns": ", ".join(df.columns),
                "Description": desc,
            }
        )
    pd.DataFrame(index_rows).to_excel(writer, sheet_name="Index", index=False)

wb = load_workbook(OUT)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="1F4E79")
body_font = Font(name="Arial")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = body_font
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)
    ws.freeze_panes = "A2"

wb.save(OUT)
print(f"Created {OUT}")
for row in index_rows:
    print(f"  {row['Sheet']}: {row['Rows']} rows")